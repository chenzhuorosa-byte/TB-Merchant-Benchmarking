"""
Converts a session's merged merchant data (listing + detail) into a
formatted Excel file. Knows nothing about Talabat field names directly —
it reads column labels from the selectors files via a passed-in map.
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

from scraper.selectors_listing import FIELDS as LISTING_FIELDS
from scraper.selectors_detail   import FIELDS as DETAIL_FIELDS

# Combined label map: field_key -> human-readable column header
ALL_LABELS = {k: v["label"] for k, v in {**LISTING_FIELDS, **DETAIL_FIELDS}.items()}

# Preferred column order: listing fields first, then detail fields
COLUMN_ORDER = list(LISTING_FIELDS.keys()) + list(DETAIL_FIELDS.keys())


def export_to_excel(session_id, session_manager):
    """
    Export all merchants for the session to an .xlsx file.
    Returns the absolute path to the created file.
    """
    rows = session_manager.get_all_merchants(session_id)
    if not rows:
        raise ValueError("No data to export for this session.")

    df = pd.DataFrame(rows)

    # Keep only columns that exist in the data, in the preferred order
    ordered_cols = [c for c in COLUMN_ORDER if c in df.columns]
    extra_cols   = [c for c in df.columns if c not in ordered_cols]
    df = df[ordered_cols + extra_cols]

    # Rename internal keys to readable headers
    df.rename(columns=ALL_LABELS, inplace=True)

    # Replace empty strings with proper NA so Excel shows blank cells
    df.replace("", pd.NA, inplace=True)

    os.makedirs("exports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join("exports", f"talabat_{session_id}_{timestamp}.xlsx")

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Merchants")
        _format_sheet(writer.sheets["Merchants"], df)

    return os.path.abspath(file_path)


def _format_sheet(ws, df):
    """Apply basic formatting: bold header, auto-width columns, freeze top row."""
    header_fill = PatternFill("solid", fgColor="1F497D")   # dark blue
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[1]:
        cell.font      = header_fill and header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit column widths (cap at 50)
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"   # freeze header row when scrolling
